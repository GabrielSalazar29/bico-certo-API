from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image, PageBreak, KeepTogether
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT, TA_JUSTIFY
from reportlab.pdfgen import canvas
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.chart import BarChart, LineChart, Reference, PieChart
from datetime import datetime
from io import BytesIO
from typing import Dict, Any, List
import matplotlib

matplotlib.use('Agg')
import matplotlib.pyplot as plt


class NumberedCanvas(canvas.Canvas):
    """Canvas personalizado com número de páginas"""

    def __init__(self, *args, **kwargs):
        canvas.Canvas.__init__(self, *args, **kwargs)
        self._saved_page_states = []

    def showPage(self):
        self._saved_page_states.append(dict(self.__dict__))
        self._startPage()

    def save(self):
        num_pages = len(self._saved_page_states)
        for state in self._saved_page_states:
            self.__dict__.update(state)
            self.draw_page_number(num_pages)
            canvas.Canvas.showPage(self)
        canvas.Canvas.save(self)

    def draw_page_number(self, page_count):
        self.setFont("Helvetica", 9)
        self.setFillColor(colors.grey)
        self.drawRightString(
            A4[0] - 0.75 * inch,
            0.5 * inch,
            f"Pagina {self._pageNumber} de {page_count}"
        )


def format_brl(amount):
    formato_us = f'{amount:,.2f}'

    return formato_us.translate(str.maketrans(',.', '.,'))


class ReportGenerator:
    """Gerador de relatórios em PDF e Excel"""

    @staticmethod
    def _create_line_chart(data: List[Dict], title: str, ylabel: str, color: str = '#4F46E5') -> BytesIO:
        """Cria gráfico de linha com proporções corretas"""
        fig, ax = plt.subplots(figsize=(10, 5))

        months = [item['month'] for item in data]
        values = [item['value'] for item in data]

        ax.plot(months, values, marker='o', linewidth=3, markersize=10,
                color=color, markerfacecolor='white', markeredgewidth=2.5, markeredgecolor=color)
        ax.fill_between(range(len(months)), values, alpha=0.2, color=color)

        ax.set_title(title, fontsize=16, fontweight='bold', pad=20)
        ax.set_ylabel(ylabel, fontsize=12, fontweight='bold')
        ax.grid(True, alpha=0.3, linestyle='--', linewidth=0.5)
        ax.spines['top'].set_visible(False)
        ax.spines['right'].set_visible(False)
        ax.spines['left'].set_linewidth(1.5)
        ax.spines['bottom'].set_linewidth(1.5)

        ax.yaxis.set_major_formatter(plt.FuncFormatter(lambda x, p: f'R$ {format_brl(x)}'))

        plt.xticks(rotation=0, fontsize=11)
        plt.yticks(fontsize=11)

        fig.tight_layout(pad=1.5)

        buffer = BytesIO()
        plt.savefig(buffer, format='png', dpi=300, bbox_inches='tight', facecolor='white')
        buffer.seek(0)
        plt.close(fig)

        return buffer

    @staticmethod
    def _create_bar_chart(data: List[Dict], title: str, xlabel: str, ylabel: str, color: str = '#10B981') -> BytesIO:
        """Cria gráfico de barras com proporções corretas"""
        fig, ax = plt.subplots(figsize=(10, 5))

        categories = [item['category'][:15] for item in data]
        values = [item.get('earnings', item.get('spent', 0)) for item in data]

        bars = ax.bar(categories, values, color=color, alpha=0.8, edgecolor='white', linewidth=2)

        for bar in bars:
            height = bar.get_height()
            if height > 0:
                ax.text(bar.get_x() + bar.get_width() / 2., height,
                        f'R$ {format_brl(height)}',
                        ha='center', va='bottom', fontsize=10, fontweight='bold')

        ax.set_title(title, fontsize=16, fontweight='bold', pad=20)
        ax.set_xlabel(xlabel, fontsize=12, fontweight='bold')
        ax.set_ylabel(ylabel, fontsize=12, fontweight='bold')
        ax.grid(True, axis='y', alpha=0.3, linestyle='--', linewidth=0.5)
        ax.spines['top'].set_visible(False)
        ax.spines['right'].set_visible(False)
        ax.spines['left'].set_linewidth(1.5)
        ax.spines['bottom'].set_linewidth(1.5)

        ax.yaxis.set_major_formatter(plt.FuncFormatter(lambda x, p: f'R$ {format_brl(x)}'))

        plt.xticks(rotation=45, ha='right', fontsize=10)
        plt.yticks(fontsize=11)

        fig.tight_layout(pad=1.5)

        buffer = BytesIO()
        plt.savefig(buffer, format='png', dpi=300, bbox_inches='tight', facecolor='white')
        buffer.seek(0)
        plt.close(fig)

        return buffer

    @staticmethod
    def _create_pie_chart(data: List[Dict], title: str) -> BytesIO:
        """Cria gráfico de pizza com proporções corretas"""
        fig, ax = plt.subplots(figsize=(8, 8))

        labels = [item['category'] for item in data[:5]]
        sizes = [item.get('earnings', item.get('spent', 0)) for item in data[:5]]

        colors_palette = ['#4F46E5', '#10B981', '#F59E0B', '#EF4444', '#8B5CF6']

        wedges, texts, autotexts = ax.pie(
            sizes,
            labels=labels,
            autopct='%1.1f%%',
            startangle=90,
            colors=colors_palette,
            wedgeprops={'edgecolor': 'white', 'linewidth': 2.5},
            textprops={'fontsize': 11, 'fontweight': 'bold'}
        )

        for autotext in autotexts:
            autotext.set_color('white')
            autotext.set_fontweight('bold')
            autotext.set_fontsize(11)

        ax.set_title(title, fontsize=16, fontweight='bold', pad=20)
        ax.axis('equal')

        fig.tight_layout(pad=1.5)

        buffer = BytesIO()
        plt.savefig(buffer, format='png', dpi=300, bbox_inches='tight', facecolor='white')
        buffer.seek(0)
        plt.close(fig)

        return buffer

    @staticmethod
    def _create_gauge_chart(value: float, max_value: float, title: str) -> BytesIO:
        """Cria gráfico de medidor estilo termômetro horizontal"""
        fig, ax = plt.subplots(figsize=(8, 3), facecolor='white')

        percentage = min(value / max_value, 1.0) if max_value > 0 else 0

        if percentage >= 0.7:
            color = '#10B981'
        elif percentage >= 0.4:
            color = '#F59E0B'
        else:
            color = '#EF4444'

        ax.barh(0, max_value, height=0.3, color='#E5E7EB', edgecolor='#CBD5E0', linewidth=2)

        if value > 0:
            ax.barh(0, value, height=0.3, color=color, edgecolor=color, linewidth=2)

        for i in [0, max_value / 4, max_value / 2, 3 * max_value / 4, max_value]:
            ax.plot([i, i], [-0.2, -0.3], 'k-', linewidth=1.5)
            ax.text(i, -0.45, f'{i:.0f}', ha='center', fontsize=10, color='#6B7280')

        ax.text(value if value > 0 else 0, 0.5, f'{value:.1f}',
                ha='center', va='bottom',
                fontsize=32, fontweight='bold',
                color=color,
                bbox=dict(boxstyle='round,pad=0.5', facecolor='white', edgecolor=color, linewidth=2))

        ax.text(max_value / 2, 0.9, title,
                ha='center', va='bottom',
                fontsize=14, fontweight='bold',
                color='#1F2937')

        ax.set_xlim(-max_value * 0.05, max_value * 1.05)
        ax.set_ylim(-0.6, 1.2)
        ax.axis('off')

        plt.tight_layout()

        buffer = BytesIO()
        plt.savefig(buffer, format='png', dpi=300, bbox_inches='tight',
                    facecolor='white', pad_inches=0.3)
        buffer.seek(0)
        plt.close(fig)

        return buffer

    @staticmethod
    def generate_provider_pdf(data: Dict[str, Any], user_name: str) -> BytesIO:
        """Gera relatório PDF do dashboard do prestador com gráficos"""
        buffer = BytesIO()
        doc = SimpleDocTemplate(
            buffer,
            pagesize=A4,
            topMargin=0.75 * inch,
            bottomMargin=0.75 * inch,
            leftMargin=0.75 * inch,
            rightMargin=0.75 * inch
        )
        elements = []

        styles = getSampleStyleSheet()

        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=28,
            textColor=colors.HexColor('#4F46E5'),
            spaceAfter=10,
            alignment=TA_CENTER,
            fontName='Helvetica-Bold'
        )

        subtitle_style = ParagraphStyle(
            'CustomSubtitle',
            parent=styles['Normal'],
            fontSize=11,
            textColor=colors.HexColor('#6B7280'),
            spaceAfter=30,
            alignment=TA_CENTER
        )

        heading_style = ParagraphStyle(
            'CustomHeading',
            parent=styles['Heading2'],
            fontSize=16,
            textColor=colors.HexColor('#1F2937'),
            spaceAfter=15,
            spaceBefore=25,
            fontName='Helvetica-Bold'
        )

        elements.append(Spacer(1, 1.5 * inch))

        elements.append(Paragraph("Dashboard do Prestador", title_style))
        elements.append(Paragraph(user_name, subtitle_style))

        info_data = [
            ['', ''],
            [Paragraph('<b>Relatorio Gerado</b>', styles['Normal']),
             Paragraph(datetime.now().strftime('%d/%m/%Y as %H:%M'), styles['Normal'])],
            [Paragraph('<b>Periodo</b>', styles['Normal']),
             Paragraph('Ultimos 6 meses', styles['Normal'])],
            ['', ''],
        ]

        info_table = Table(info_data, colWidths=[2 * inch, 3 * inch])
        info_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor('#F9FAFB')),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('FONTNAME', (0, 1), (0, -2), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('TOPPADDING', (0, 0), (-1, -1), 12),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 12),
            ('BOX', (0, 0), (-1, -1), 2, colors.HexColor('#E5E7EB')),
            ('LINEBELOW', (0, 1), (-1, -2), 1, colors.HexColor('#E5E7EB')),
        ]))

        elements.append(info_table)
        elements.append(PageBreak())

        elements.append(Spacer(1, 1.5 * inch))
        elements.append(Paragraph("Resumo Executivo", heading_style))

        trends = data.get('trends', {})
        metrics = data.get('metrics', {})

        card1_data = [
            ['Jobs Concluidos'],
            [str(data.get('completedJobs', 0))],
        ]
        card1 = Table(card1_data, colWidths=[3 * inch], rowHeights=[0.5 * inch, 1 * inch])
        card1.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor('#F0FDF4')),
            ('BOX', (0, 0), (-1, -1), 3, colors.HexColor('#10B981')),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('VALIGN', (0, 0), (0, 0), 'BOTTOM'),
            ('VALIGN', (0, 1), (0, 1), 'TOP'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 13),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.HexColor('#6B7280')),
            ('FONTNAME', (0, 1), (-1, 1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 1), (-1, 1), 32),
            ('TEXTCOLOR', (0, 1), (-1, 1), colors.HexColor('#10B981')),
            ('TOPPADDING', (0, 0), (-1, 0), 15),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 2),
            ('TOPPADDING', (0, 1), (-1, 1), 5),
            ('BOTTOMPADDING', (0, 1), (-1, 1), 20),
        ]))

        card2_data = [
            ['Ganhos Totais'],
            [f'R$ {format_brl(data.get("totalEarnings", 0))}'],
        ]
        card2 = Table(card2_data, colWidths=[3 * inch], rowHeights=[0.5 * inch, 1 * inch])
        card2.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor('#EEF2FF')),
            ('BOX', (0, 0), (-1, -1), 3, colors.HexColor('#4F46E5')),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('VALIGN', (0, 0), (0, 0), 'BOTTOM'),
            ('VALIGN', (0, 1), (0, 1), 'TOP'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 13),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.HexColor('#6B7280')),
            ('FONTNAME', (0, 1), (-1, 1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 1), (-1, 1), 28),
            ('TEXTCOLOR', (0, 1), (-1, 1), colors.HexColor('#4F46E5')),
            ('TOPPADDING', (0, 0), (-1, 0), 15),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 2),
            ('TOPPADDING', (0, 1), (-1, 1), 5),
            ('BOTTOMPADDING', (0, 1), (-1, 1), 20),
        ]))

        row1 = Table([[card1, card2]], colWidths=[3.25 * inch, 3.25 * inch])
        row1.setStyle(TableStyle([
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ]))
        elements.append(row1)
        elements.append(Spacer(1, 12))

        card3_data = [
            ['Avaliacao Media'],
            [f'{data.get("averageRating", 0):.1f}'],
        ]
        card3 = Table(card3_data, colWidths=[3 * inch], rowHeights=[0.5 * inch, 1 * inch])
        card3.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor('#FFFBEB')),
            ('BOX', (0, 0), (-1, -1), 3, colors.HexColor('#F59E0B')),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('VALIGN', (0, 0), (0, 0), 'BOTTOM'),
            ('VALIGN', (0, 1), (0, 1), 'TOP'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 13),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.HexColor('#6B7280')),
            ('FONTNAME', (0, 1), (-1, 1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 1), (-1, 1), 32),
            ('TEXTCOLOR', (0, 1), (-1, 1), colors.HexColor('#F59E0B')),
            ('TOPPADDING', (0, 0), (-1, 0), 15),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 2),
            ('TOPPADDING', (0, 1), (-1, 1), 5),
            ('BOTTOMPADDING', (0, 1), (-1, 1), 20),
        ]))

        card4_data = [
            ['Taxa de Aceitacao'],
            [f'{data.get("proposalAcceptanceRate", 0)}%'],
        ]
        card4 = Table(card4_data, colWidths=[3 * inch], rowHeights=[0.5 * inch, 1 * inch])
        card4.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor('#FAF5FF')),
            ('BOX', (0, 0), (-1, -1), 3, colors.HexColor('#8B5CF6')),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('VALIGN', (0, 0), (0, 0), 'BOTTOM'),
            ('VALIGN', (0, 1), (0, 1), 'TOP'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 13),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.HexColor('#6B7280')),
            ('FONTNAME', (0, 1), (-1, 1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 1), (-1, 1), 32),
            ('TEXTCOLOR', (0, 1), (-1, 1), colors.HexColor('#8B5CF6')),
            ('TOPPADDING', (0, 0), (-1, 0), 15),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 2),
            ('TOPPADDING', (0, 1), (-1, 1), 5),
            ('BOTTOMPADDING', (0, 1), (-1, 1), 20),
        ]))

        row2 = Table([[card3, card4]], colWidths=[3.25 * inch, 3.25 * inch])
        row2.setStyle(TableStyle([
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ]))
        elements.append(row2)
        elements.append(Spacer(1, 20))

        trend_data = [
            ['Metrica', 'Tendencia (30 dias)'],
            ['Crescimento de Jobs', '{}{}%'.format(
                '+' if trends.get('jobsTrend', 0) >= 0 else '',
                trends.get('jobsTrend', 0))],
            ['Crescimento de Ganhos', '{}{}%'.format(
                '+' if trends.get('earningsTrend', 0) >= 0 else '',
                trends.get('earningsTrend', 0))],
        ]

        trend_table = Table(trend_data, colWidths=[3.25 * inch, 3.25 * inch])
        trend_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1F2937')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('TOPPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#F9FAFB')),
            ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#E5E7EB')),
            ('FONTSIZE', (0, 1), (-1, -1), 11),
            ('TOPPADDING', (0, 1), (-1, -1), 10),
            ('BOTTOMPADDING', (0, 1), (-1, -1), 10),
        ]))

        elements.append(trend_table)
        elements.append(PageBreak())

        elements.append(Spacer(1, 1.5 * inch))

        elements.append(Paragraph("Evolucao de Ganhos", heading_style))

        monthly_earnings = data.get('monthlyEarnings', [])
        if monthly_earnings:
            chart_buffer = ReportGenerator._create_line_chart(
                monthly_earnings,
                'Ganhos Mensais (Ultimos 6 meses)',
                'Ganhos (R$)',
                '#4F46E5'
            )
            chart_img = Image(chart_buffer, width=6.5 * inch, height=3.25 * inch)
            elements.append(chart_img)
        else:
            elements.append(Paragraph(
                '<i>Nenhum dado de ganhos disponivel</i>',
                ParagraphStyle('Italic', parent=styles['Normal'], textColor=colors.grey)
            ))

        elements.append(Spacer(1, 20))

        total_earnings = data.get('totalEarnings', 0)
        avg_monthly = total_earnings / 6 if total_earnings > 0 else 0

        earnings_stats = [
            ['Total Acumulado', 'Media Mensal', 'Maior Ganho'],
            [f'R$ {format_brl(total_earnings)}',
             f'R$ {format_brl(avg_monthly)}',
             f'R$ {format_brl(metrics.get("highestEarningJob", 0))}'],
        ]

        earnings_table = Table(earnings_stats, colWidths=[2.16 * inch, 2.16 * inch, 2.16 * inch])
        earnings_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4F46E5')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 11),
            ('TOPPADDING', (0, 0), (-1, -1), 12),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#EEF2FF')),
            ('FONTSIZE', (0, 1), (-1, -1), 14),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica-Bold'),
            ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#4F46E5')),
        ]))

        elements.append(earnings_table)
        elements.append(Spacer(1, 1.5 * inch))
        elements.append(PageBreak())

        elements.append(Paragraph("Performance por Categoria", heading_style))

        categories = data.get('jobsByCategory', [])
        if categories:
            chart_buffer = ReportGenerator._create_bar_chart(
                categories,
                'Ganhos por Categoria',
                'Categoria',
                'Ganhos (R$)',
                '#10B981'
            )
            chart_img = Image(chart_buffer, width=6.5 * inch, height=3.25 * inch)
            elements.append(chart_img)
            elements.append(Spacer(1, 40))

            pie_buffer = ReportGenerator._create_pie_chart(
                categories,
                'Distribuicao de Ganhos por Categoria'
            )
            pie_img = Image(pie_buffer, width=5 * inch, height=5 * inch)

            pie_container = Table([[pie_img]], colWidths=[6.5 * inch])
            pie_container.setStyle(TableStyle([
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ]))

            elements.append(pie_container)
            elements.append(PageBreak())

            cat_data = [['Categoria', 'Jobs', 'Ganhos', 'Media por Job']]
            for cat in categories:
                avg_per_job = cat['earnings'] / cat['count'] if cat['count'] > 0 else 0
                cat_data.append([
                    cat['category'],
                    str(cat['count']),
                    f"R$ {format_brl(cat['earnings'])}",
                    f"R$ {format_brl(avg_per_job)}"
                ])

            cat_table = Table(cat_data, colWidths=[2 * inch, 1.5 * inch, 1.5 * inch, 1.5 * inch])
            cat_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#10B981')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 11),
                ('TOPPADDING', (0, 0), (-1, -1), 10),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 10),
                ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#F0FDF4')),
                ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#10B981')),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F0FDF4')]),
            ]))

            elements.append(cat_table)
        else:
            elements.append(Paragraph(
                '<i>Nenhuma categoria registrada</i>',
                ParagraphStyle('Italic', parent=styles['Normal'], textColor=colors.grey)
            ))

        elements.append(Paragraph("Metricas Detalhadas", heading_style))

        metrics_data = [
            ['Metrica', 'Valor'],
            ['Valor Medio por Job', f'R$ {format_brl(metrics.get("averageJobValue", 0))}'],
            ['Tempo Medio de Entrega', f'{metrics.get("averageDeliveryTime", 0):.1f} dias'],
            ['Total de Clientes Atendidos', str(metrics.get("totalClients", 0))],
            ['Job Mais Lucrativo', f'R$ {format_brl(metrics.get("highestEarningJob", 0))}'],
            ['Ultimo Job Realizado', metrics.get("lastJobDate", "N/A")],
            ['Jobs Ativos', str(data.get("activeJobs", 0))],
            ['Propostas Pendentes', str(data.get("pendingProposals", 0))],
        ]

        metrics_table = Table(metrics_data, colWidths=[3.5 * inch, 3 * inch])
        metrics_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#8B5CF6')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 11),
            ('TOPPADDING', (0, 0), (-1, -1), 12),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#FAF5FF')),
            ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#8B5CF6')),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#FAF5FF')]),
            ('FONTSIZE', (1, 1), (1, -1), 12),
            ('FONTNAME', (1, 1), (1, -1), 'Helvetica-Bold'),
        ]))

        elements.append(metrics_table)
        elements.append(Spacer(1, 30))

        footer_style = ParagraphStyle(
            'Footer',
            parent=styles['Normal'],
            fontSize=9,
            textColor=colors.HexColor('#6B7280'),
            alignment=TA_CENTER,
            borderColor=colors.HexColor('#E5E7EB'),
            borderWidth=1,
            borderPadding=10
        )

        footer_text = f"""
        <b>Bico Certo - Plataforma de Servicos Descentralizada</b><br/>
        Relatorio gerado automaticamente em {datetime.now().strftime('%d/%m/%Y as %H:%M')}<br/>
        """

        elements.append(Paragraph(footer_text, footer_style))

        doc.build(elements, canvasmaker=NumberedCanvas)
        buffer.seek(0)
        return buffer

    @staticmethod
    def generate_client_pdf(data: Dict[str, Any], user_name: str) -> BytesIO:
        """Gera relatório PDF do dashboard do cliente com gráficos"""
        buffer = BytesIO()
        doc = SimpleDocTemplate(
            buffer,
            pagesize=A4,
            topMargin=0.75 * inch,
            bottomMargin=0.75 * inch,
            leftMargin=0.75 * inch,
            rightMargin=0.75 * inch
        )
        elements = []

        styles = getSampleStyleSheet()

        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=28,
            textColor=colors.HexColor('#059669'),
            spaceAfter=10,
            alignment=TA_CENTER,
            fontName='Helvetica-Bold'
        )

        subtitle_style = ParagraphStyle(
            'CustomSubtitle',
            parent=styles['Normal'],
            fontSize=11,
            textColor=colors.HexColor('#6B7280'),
            spaceAfter=30,
            alignment=TA_CENTER
        )

        heading_style = ParagraphStyle(
            'CustomHeading',
            parent=styles['Heading2'],
            fontSize=16,
            textColor=colors.HexColor('#1F2937'),
            spaceAfter=15,
            spaceBefore=25,
            fontName='Helvetica-Bold'
        )

        elements.append(Spacer(1, 1.5 * inch))
        elements.append(Paragraph("Dashboard do Cliente", title_style))
        elements.append(Paragraph(user_name, subtitle_style))

        info_data = [
            ['', ''],
            [Paragraph('<b>Relatorio Gerado</b>', styles['Normal']),
             Paragraph(datetime.now().strftime('%d/%m/%Y as %H:%M'), styles['Normal'])],
            [Paragraph('<b>Periodo</b>', styles['Normal']),
             Paragraph('Ultimos 6 meses', styles['Normal'])],
            ['', ''],
        ]

        info_table = Table(info_data, colWidths=[2 * inch, 3 * inch])
        info_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor('#F0FDF4')),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('FONTNAME', (0, 1), (0, -2), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('TOPPADDING', (0, 0), (-1, -1), 12),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 12),
            ('BOX', (0, 0), (-1, -1), 2, colors.HexColor('#059669')),
            ('LINEBELOW', (0, 1), (-1, -2), 1, colors.HexColor('#D1FAE5')),
        ]))

        elements.append(info_table)
        elements.append(PageBreak())

        elements.append(Spacer(1, 2 * inch))
        elements.append(Paragraph("Resumo Executivo", heading_style))

        card1_data = [
            ['Jobs Ativos'],
            [str(data.get('activeJobs', 0))],
        ]
        card1 = Table(card1_data, colWidths=[3 * inch], rowHeights=[0.5 * inch, 1 * inch])
        card1.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor('#EFF6FF')),
            ('BOX', (0, 0), (-1, -1), 3, colors.HexColor('#3B82F6')),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('VALIGN', (0, 0), (0, 0), 'BOTTOM'),
            ('VALIGN', (0, 1), (0, 1), 'TOP'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 13),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.HexColor('#6B7280')),
            ('FONTNAME', (0, 1), (-1, 1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 1), (-1, 1), 32),
            ('TEXTCOLOR', (0, 1), (-1, 1), colors.HexColor('#3B82F6')),
            ('TOPPADDING', (0, 0), (-1, 0), 15),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 2),
            ('TOPPADDING', (0, 1), (-1, 1), 5),
            ('BOTTOMPADDING', (0, 1), (-1, 1), 20),
        ]))

        card2_data = [
            ['Jobs Concluidos'],
            [str(data.get('completedJobs', 0))],
        ]
        card2 = Table(card2_data, colWidths=[3 * inch], rowHeights=[0.5 * inch, 1 * inch])
        card2.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor('#F0FDF4')),
            ('BOX', (0, 0), (-1, -1), 3, colors.HexColor('#10B981')),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('VALIGN', (0, 0), (0, 0), 'BOTTOM'),
            ('VALIGN', (0, 1), (0, 1), 'TOP'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 13),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.HexColor('#6B7280')),
            ('FONTNAME', (0, 1), (-1, 1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 1), (-1, 1), 32),
            ('TEXTCOLOR', (0, 1), (-1, 1), colors.HexColor('#10B981')),
            ('TOPPADDING', (0, 0), (-1, 0), 15),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 2),
            ('TOPPADDING', (0, 1), (-1, 1), 5),
            ('BOTTOMPADDING', (0, 1), (-1, 1), 20),
        ]))

        row1 = Table([[card1, card2]], colWidths=[3.25 * inch, 3.25 * inch])
        row1.setStyle(TableStyle([
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ]))
        elements.append(row1)
        elements.append(Spacer(1, 12))

        card3_data = [
            ['Total Gasto'],
            [f'R$ {format_brl(data.get("totalSpent", 0))}'],
        ]
        card3 = Table(card3_data, colWidths=[3 * inch], rowHeights=[0.5 * inch, 1 * inch])
        card3.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor('#FAF5FF')),
            ('BOX', (0, 0), (-1, -1), 3, colors.HexColor('#8B5CF6')),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('VALIGN', (0, 0), (0, 0), 'BOTTOM'),
            ('VALIGN', (0, 1), (0, 1), 'TOP'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 13),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.HexColor('#6B7280')),
            ('FONTNAME', (0, 1), (-1, 1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 1), (-1, 1), 28),
            ('TEXTCOLOR', (0, 1), (-1, 1), colors.HexColor('#8B5CF6')),
            ('TOPPADDING', (0, 0), (-1, 0), 15),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 2),
            ('TOPPADDING', (0, 1), (-1, 1), 5),
            ('BOTTOMPADDING', (0, 1), (-1, 1), 20),
        ]))

        card4_data = [
            ['Prestadores'],
            [str(data.get('providersHired', 0))],
        ]
        card4 = Table(card4_data, colWidths=[3 * inch], rowHeights=[0.5 * inch, 1 * inch])
        card4.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor('#FFFBEB')),
            ('BOX', (0, 0), (-1, -1), 3, colors.HexColor('#F59E0B')),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('VALIGN', (0, 0), (0, 0), 'BOTTOM'),
            ('VALIGN', (0, 1), (0, 1), 'TOP'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 13),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.HexColor('#6B7280')),
            ('FONTNAME', (0, 1), (-1, 1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 1), (-1, 1), 32),
            ('TEXTCOLOR', (0, 1), (-1, 1), colors.HexColor('#F59E0B')),
            ('TOPPADDING', (0, 0), (-1, 0), 15),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 2),
            ('TOPPADDING', (0, 1), (-1, 1), 5),
            ('BOTTOMPADDING', (0, 1), (-1, 1), 20),
        ]))

        row2 = Table([[card3, card4]], colWidths=[3.25 * inch, 3.25 * inch])
        row2.setStyle(TableStyle([
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ]))
        elements.append(row2)
        elements.append(Spacer(1, 20))

        elements.append(PageBreak())

        elements.append(Spacer(1, 2 * inch))

        elements.append(Paragraph("Evolucao de Gastos", heading_style))

        monthly_spending = data.get('monthlySpending', [])
        if monthly_spending:
            chart_buffer = ReportGenerator._create_line_chart(
                monthly_spending,
                'Gastos Mensais (Ultimos 6 meses)',
                'Gastos (R$)',
                '#059669'
            )
            chart_img = Image(chart_buffer, width=6.5 * inch, height=3.25 * inch)
            elements.append(chart_img)

        elements.append(Spacer(1, 2 * inch))

        elements.append(PageBreak())

        elements.append(Paragraph("Gastos por Categoria", heading_style))

        categories = data.get('spendingByCategory', [])
        if categories:
            chart_buffer = ReportGenerator._create_bar_chart(
                categories,
                'Gastos por Categoria',
                'Categoria',
                'Gastos (R$)',
                '#8B5CF6'
            )
            chart_img = Image(chart_buffer, width=6.5 * inch, height=3.25 * inch)
            elements.append(chart_img)
            elements.append(Spacer(1, 40))

            pie_buffer = ReportGenerator._create_pie_chart(
                categories,
                'Distribuicao de Gastos por Categoria'
            )
            pie_img = Image(pie_buffer, width=5 * inch, height=5 * inch)

            pie_container = Table([[pie_img]], colWidths=[6.5 * inch])
            pie_container.setStyle(TableStyle([
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ]))

            elements.append(pie_container)
            elements.append(Spacer(1, 1 * inch))

        elements.append(Spacer(1, 1.5 * inch))

        elements.append(Paragraph("Estatisticas Gerais", heading_style))

        metrics = data.get('metrics', {})
        elements.append(Spacer(1, 20))

        metrics_data = [
            ['Metrica', 'Valor'],
            ['Taxa de Conclusao', f'{metrics.get("completionRate", 0)}%'],
            ['Categoria Favorita', metrics.get("favoriteCategory", "N/A")],
            ['Aprovacoes Pendentes', str(data.get("pendingApprovals", 0))],
        ]

        metrics_table = Table(metrics_data, colWidths=[3.5 * inch, 3 * inch])
        metrics_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#059669')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 11),
            ('TOPPADDING', (0, 0), (-1, -1), 12),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#F0FDF4')),
            ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#059669')),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F0FDF4')]),
            ('FONTSIZE', (1, 1), (1, -1), 12),
            ('FONTNAME', (1, 1), (1, -1), 'Helvetica-Bold'),
        ]))

        elements.append(metrics_table)
        elements.append(Spacer(1, 30))

        footer_style = ParagraphStyle(
            'Footer',
            parent=styles['Normal'],
            fontSize=9,
            textColor=colors.HexColor('#6B7280'),
            alignment=TA_CENTER,
            borderColor=colors.HexColor('#059669'),
            borderWidth=1,
            borderPadding=10
        )

        footer_text = f"""
        <b>Bico Certo - Plataforma de Servicos Descentralizada</b><br/>
        Relatorio gerado automaticamente em {datetime.now().strftime('%d/%m/%Y as %H:%M')}<br/>
        """

        elements.append(Paragraph(footer_text, footer_style))

        doc.build(elements, canvasmaker=NumberedCanvas)
        buffer.seek(0)
        return buffer

    @staticmethod
    def generate_provider_excel(data: Dict[str, Any], user_name: str) -> BytesIO:
        """Gera relatório Excel do dashboard do prestador"""
        buffer = BytesIO()
        wb = Workbook()

        header_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        ws1 = wb.active
        ws1.title = "Resumo"

        ws1['A1'] = f"Dashboard do Prestador - {user_name}"
        ws1['A1'].font = Font(bold=True, size=16, color="4F46E5")
        ws1['A2'] = f"Gerado em: {datetime.now().strftime('%d/%m/%Y as %H:%M')}"

        trends = data.get('trends', {})

        ws1['A4'] = 'Metrica'
        ws1['B4'] = 'Valor'
        ws1['A4'].fill = header_fill
        ws1['B4'].fill = header_fill
        ws1['A4'].font = header_font
        ws1['B4'].font = header_font

        metrics_summary = [
            ['Jobs Concluidos', data.get('completedJobs', 0)],
            ['Ganhos Totais', f"R$ {format_brl(data.get('totalEarnings', 0))}"],
            ['Avaliacao Media', f"{data.get('averageRating', 0):.1f}"],
            ['Taxa de Aceitacao', f"{data.get('proposalAcceptanceRate', 0)}%"],
            ['Jobs Ativos', data.get('activeJobs', 0)],
            ['Propostas Pendentes', data.get('pendingProposals', 0)],
            ['Tendencia de Jobs', f"{trends.get('jobsTrend', 0):+.1f}%"],
            ['Tendencia de Ganhos', f"{trends.get('earningsTrend', 0):+.1f}%"],
        ]

        for idx, (metric, value) in enumerate(metrics_summary, start=5):
            ws1[f'A{idx}'] = metric
            ws1[f'B{idx}'] = value
            ws1[f'A{idx}'].border = border
            ws1[f'B{idx}'].border = border

        ws1.column_dimensions['A'].width = 30
        ws1.column_dimensions['B'].width = 20

        ws2 = wb.create_sheet("Categorias")
        ws2['A1'] = 'Performance por Categoria'
        ws2['A1'].font = Font(bold=True, size=14, color="10B981")

        ws2['A3'] = 'Categoria'
        ws2['B3'] = 'Quantidade de Jobs'
        ws2['C3'] = 'Ganhos Totais'

        for col in ['A3', 'B3', 'C3']:
            ws2[col].fill = PatternFill(start_color="10B981", end_color="10B981", fill_type="solid")
            ws2[col].font = header_font

        categories = data.get('jobsByCategory', [])
        for idx, cat in enumerate(categories, start=4):
            ws2[f'A{idx}'] = cat['category']
            ws2[f'B{idx}'] = cat['count']
            ws2[f'C{idx}'] = cat['earnings']

            for col in ['A', 'B', 'C']:
                ws2[f'{col}{idx}'].border = border

        ws2.column_dimensions['A'].width = 25
        ws2.column_dimensions['B'].width = 20
        ws2.column_dimensions['C'].width = 20

        ws3 = wb.create_sheet("Ganhos Mensais")
        ws3['A1'] = 'Evolucao de Ganhos'
        ws3['A1'].font = Font(bold=True, size=14, color="4F46E5")

        ws3['A3'] = 'Mes'
        ws3['B3'] = 'Valor (R$)'

        for col in ['A3', 'B3']:
            ws3[col].fill = header_fill
            ws3[col].font = header_font

        monthly = data.get('monthlyEarnings', [])
        for idx, month_data in enumerate(monthly, start=4):
            ws3[f'A{idx}'] = month_data['month']
            ws3[f'B{idx}'] = month_data['value']

            for col in ['A', 'B']:
                ws3[f'{col}{idx}'].border = border

        if monthly:
            chart = LineChart()
            chart.title = "Evolucao de Ganhos"
            chart.style = 10
            chart.y_axis.title = 'Ganhos (R$)'
            chart.x_axis.title = 'Mes'

            data_ref = Reference(ws3, min_col=2, min_row=3, max_row=len(monthly) + 3)
            cats = Reference(ws3, min_col=1, min_row=4, max_row=len(monthly) + 3)
            chart.add_data(data_ref, titles_from_data=True)
            chart.set_categories(cats)

            ws3.add_chart(chart, "D3")

        ws3.column_dimensions['A'].width = 15
        ws3.column_dimensions['B'].width = 20

        wb.save(buffer)
        buffer.seek(0)
        return buffer

    @staticmethod
    def generate_client_excel(data: Dict[str, Any], user_name: str) -> BytesIO:
        """Gera relatório Excel do dashboard do cliente"""
        buffer = BytesIO()
        wb = Workbook()

        header_fill = PatternFill(start_color="059669", end_color="059669", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        ws1 = wb.active
        ws1.title = "Resumo"

        ws1['A1'] = f"Dashboard do Cliente - {user_name}"
        ws1['A1'].font = Font(bold=True, size=16, color="059669")
        ws1['A2'] = f"Gerado em: {datetime.now().strftime('%d/%m/%Y as %H:%M')}"

        ws1['A4'] = 'Metrica'
        ws1['B4'] = 'Valor'
        ws1['A4'].fill = header_fill
        ws1['B4'].fill = header_fill
        ws1['A4'].font = header_font
        ws1['B4'].font = header_font

        metrics_summary = [
            ['Jobs Ativos', data.get('activeJobs', 0)],
            ['Jobs Concluidos', data.get('completedJobs', 0)],
            ['Total Gasto', f"R$ {format_brl(data.get('totalSpent', 0))}"],
            ['Prestadores Contratados', data.get('providersHired', 0)],
            ['Aprovacoes Pendentes', data.get('pendingApprovals', 0)],
        ]

        for idx, (metric, value) in enumerate(metrics_summary, start=5):
            ws1[f'A{idx}'] = metric
            ws1[f'B{idx}'] = value
            ws1[f'A{idx}'].border = border
            ws1[f'B{idx}'].border = border

        ws1.column_dimensions['A'].width = 30
        ws1.column_dimensions['B'].width = 20

        ws2 = wb.create_sheet("Categorias")
        ws2['A1'] = 'Gastos por Categoria'
        ws2['A1'].font = Font(bold=True, size=14, color="8B5CF6")

        ws2['A3'] = 'Categoria'
        ws2['B3'] = 'Quantidade'
        ws2['C3'] = 'Valor Gasto'

        for col in ['A3', 'B3', 'C3']:
            ws2[col].fill = PatternFill(start_color="8B5CF6", end_color="8B5CF6", fill_type="solid")
            ws2[col].font = header_font

        categories = data.get('spendingByCategory', [])
        for idx, cat in enumerate(categories, start=4):
            ws2[f'A{idx}'] = cat['category']
            ws2[f'B{idx}'] = cat['count']
            ws2[f'C{idx}'] = cat['spent']

            for col in ['A', 'B', 'C']:
                ws2[f'{col}{idx}'].border = border

        ws2.column_dimensions['A'].width = 25
        ws2.column_dimensions['B'].width = 15
        ws2.column_dimensions['C'].width = 20

        ws3 = wb.create_sheet("Gastos Mensais")
        ws3['A1'] = 'Evolucao de Gastos'
        ws3['A1'].font = Font(bold=True, size=14, color="059669")

        ws3['A3'] = 'Mes'
        ws3['B3'] = 'Valor (R$)'

        for col in ['A3', 'B3']:
            ws3[col].fill = header_fill
            ws3[col].font = header_font

        monthly = data.get('monthlySpending', [])
        for idx, month_data in enumerate(monthly, start=4):
            ws3[f'A{idx}'] = month_data['month']
            ws3[f'B{idx}'] = month_data['value']

            for col in ['A', 'B']:
                ws3[f'{col}{idx}'].border = border

        if monthly:
            chart = LineChart()
            chart.title = "Evolucao de Gastos"
            chart.style = 10
            chart.y_axis.title = 'Gastos (R$)'
            chart.x_axis.title = 'Mes'

            data_ref = Reference(ws3, min_col=2, min_row=3, max_row=len(monthly) + 3)
            cats = Reference(ws3, min_col=1, min_row=4, max_row=len(monthly) + 3)
            chart.add_data(data_ref, titles_from_data=True)
            chart.set_categories(cats)

            ws3.add_chart(chart, "D3")

        ws3.column_dimensions['A'].width = 15
        ws3.column_dimensions['B'].width = 20

        wb.save(buffer)
        buffer.seek(0)
        return buffer